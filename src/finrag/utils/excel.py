import io
from typing import Any, Dict, List

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = structlog.get_logger(__name__)


def generate_financial_excel(sheets_data: Dict[str, List[List[Any]]]) -> bytes:
    """Generate a premium Excel workbook with active mathematical formulas.

    Args:
        sheets_data: A mapping of sheet names to grids of data (rows of columns).

    Returns:
        Bytes representing the compiled Excel workbook.
    """
    logger.info("Generating financial Excel workbook", sheets=list(sheets_data.keys()))
    wb = Workbook()

    # Remove the default sheet created by openpyxl
    if wb.active:
        wb.remove(wb.active)

    # Style definitions
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Navy blue
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    double_border_side = Side(border_style="double", color="333333")

    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_border_side, bottom=double_border_side)

    for sheet_name, grid in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        # Keep track of label rows to construct formulas
        label_to_row: Dict[str, int] = {}

        for row_idx, row in enumerate(grid, start=1):
            # Record label from Column A
            if len(row) > 0 and isinstance(row[0], str):
                label_clean = row[0].strip().lower()
                label_to_row[label_clean] = row_idx

            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_regular
                cell.border = border_data

                col_letter = get_column_letter(col_idx)

                # Flag if this is a total or margin summary row
                is_summary = False
                row_label = str(row[0]).strip().lower() if len(row) > 0 else ""
                if any(x in row_label for x in ["total", "profit", "margin", "income", "change"]):
                    is_summary = True

                # Header styling
                if row_idx == 1:
                    cell.value = val
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    continue

                # Auto-inject formulas for subsequent data columns (Column B onwards)
                if col_idx > 1 and isinstance(val, (int, float, str)):
                    # Check for formulas based on labels
                    revenue_row = label_to_row.get("total revenue") or label_to_row.get("revenue")
                    cogs_row = label_to_row.get("cost of goods sold") or label_to_row.get("cost of revenue")
                    gp_row = label_to_row.get("gross profit")
                    opex_row = label_to_row.get("operating expenses") or label_to_row.get("total operating expenses")
                    opinc_row = label_to_row.get("operating income")
                    netinc_row = label_to_row.get("net income")

                    formula_set = False

                    if row_label == "gross profit" and revenue_row and cogs_row:
                        cell.value = f"={col_letter}{revenue_row}-{col_letter}{cogs_row}"
                        formula_set = True
                    elif row_label == "operating income" and gp_row and opex_row:
                        cell.value = f"={col_letter}{gp_row}-{col_letter}{opex_row}"
                        formula_set = True
                    elif row_label == "gross margin" and gp_row and revenue_row:
                        cell.value = f"={col_letter}{gp_row}/{col_letter}{revenue_row}"
                        cell.number_format = "0.0%"
                        formula_set = True
                    elif row_label == "operating margin" and opinc_row and revenue_row:
                        cell.value = f"={col_letter}{opinc_row}/{col_letter}{revenue_row}"
                        cell.number_format = "0.0%"
                        formula_set = True
                    elif row_label == "net margin" and netinc_row and revenue_row:
                        cell.value = f"={col_letter}{netinc_row}/{col_letter}{revenue_row}"
                        cell.number_format = "0.0%"
                        formula_set = True
                    elif "margin" in row_label and isinstance(val, float) and val <= 1.0:
                        cell.value = val
                        cell.number_format = "0.0%"
                        formula_set = True

                    if not formula_set:
                        # Write flat value if no formula is matched
                        cell.value = val
                        if isinstance(val, (int, float)):
                            cell.number_format = "#,##0"

                else:
                    # Column A (labels) or standard strings
                    cell.value = val

                # Apply styling accents for summary rows
                if is_summary:
                    cell.font = font_bold
                    cell.fill = fill_total
                    cell.border = border_total
                    if col_idx > 1:
                        cell.alignment = Alignment(horizontal="right")
                else:
                    if col_idx > 1:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column width parameters
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    logger.info("Compiled Excel workbook successfully", bytes_count=len(buf.getvalue()))
    return buf.getvalue()
