import io
import pytest
from unittest.mock import MagicMock
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from finrag.core.security import create_access_token
from finrag.utils.excel import generate_financial_excel
from finrag.utils.pdf import compile_pdf_report


@pytest.fixture
def auth_header() -> dict:
    """Provide a valid mock JWT authorization header with queries read scopes."""
    token = create_access_token(
        data={"sub": "test_user", "scopes": ["read:queries"]}
    )
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def unauthorized_header() -> dict:
    """Provide a JWT token lacking read:queries scope."""
    token = create_access_token(
        data={"sub": "test_user", "scopes": ["read:documents"]}
    )
    return {"Authorization": f"Bearer {token}"}


# --- Excel Compiler Tests ---

def test_excel_formula_generation() -> None:
    """Verify excel builder inserts active mathematical cell formulas."""
    sheets_data = {
        "Income Statement": [
            ["Item", "Q1 2025", "Q2 2025"],
            ["Total Revenue", 1000, 1200],
            ["Cost of Goods Sold", 400, 450],
            ["Gross Profit", 0, 0],  # Mock values, will be formula
            ["Gross Margin", 0.0, 0.0],
        ]
    }

    excel_bytes = generate_financial_excel(sheets_data)
    assert len(excel_bytes) > 0

    # Load workbook using openpyxl and assert formulas are set
    wb = load_workbook(io.BytesIO(excel_bytes))
    sheet = wb["Income Statement"]

    # Row 4 is "Gross Profit"
    # Column B is Q1 2025 (col 2). Row 2 is Total Revenue, Row 3 is Cost of Goods Sold
    # GP Formula should be =B2-B3
    assert sheet.cell(row=4, column=2).value == "=B2-B3"
    assert sheet.cell(row=4, column=3).value == "=C2-C3"

    # Row 5 is "Gross Margin". Formula should be =B4/B2
    assert sheet.cell(row=5, column=2).value == "=B4/B2"
    assert sheet.cell(row=5, column=3).value == "=C4/C2"


# --- PDF Compiler Tests ---

def test_pdf_appendix_generation() -> None:
    """Verify PDF compiler aggregates report body and coordinate citation tables."""
    answer = "Operating margin declined due to rising opex [1]."
    citations = [
        {
            "chunk_id": "c1",
            "document_id": "doc-uuid-1234",
            "page": 12,
            "bbox": {"x1": 100, "y1": 200, "x2": 800, "y2": 300},
            "section": "Management Discussion",
        }
    ]

    pdf_bytes = compile_pdf_report(answer, citations, title="Q2 Analytics Summary")
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b"%PDF")  # PDF magic number header


# --- API Routes Tests ---

def test_api_export_pdf_success(client: TestClient, auth_header: dict) -> None:
    """Verify API POST /export/pdf endpoint returns PDF binary response."""
    payload = {
        "answer": "Net revenues increased by 15% YoY.",
        "citations": [
            {
                "chunk_id": "c1",
                "document_id": "doc-uuid-1234",
                "page": 12,
                "bbox": {"x1": 100, "y1": 200, "x2": 800, "y2": 300},
                "section": "Management Discussion",
            }
        ],
        "title": "Quarterly Report Compile"
    }

    response = client.post(
        "/api/v1/queries/export/pdf",
        json=payload,
        headers=auth_header
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert "attachment; filename=report.pdf" in response.headers["content-disposition"]
    assert len(response.content) > 0


def test_api_export_excel_success(client: TestClient, auth_header: dict) -> None:
    """Verify API POST /export/excel endpoint returns Excel spreadsheet response."""
    payload = {
        "sheets": {
            "Balance Sheet": [
                ["Metric", "Value"],
                ["Total Assets", 50000],
                ["Total Liabilities", 30000],
            ]
        }
    }

    response = client.post(
        "/api/v1/queries/export/excel",
        json=payload,
        headers=auth_header
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=financials.xlsx" in response.headers["content-disposition"]
    assert len(response.content) > 0


def test_api_export_insufficient_scope(client: TestClient, unauthorized_header: dict) -> None:
    """Verify RBAC scopes verify queries permissions."""
    payload = {"answer": "Some analysis text", "citations": []}
    response = client.post(
        "/api/v1/queries/export/pdf",
        json=payload,
        headers=unauthorized_header
    )
    assert response.status_code == 403
